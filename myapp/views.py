from django.shortcuts import render
import openpyxl
import pandas as pd
from datetime import datetime
import os,time

def index(request):
    file_path = 'myapp/templates/myapp/result.xlsx'

    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        # print(sheets)

        # getting a particular sheet
        try:
            worksheet = wb["Sheet1"]
        except:
            worksheet = wb["Лист1"]
        # print(type(worksheet))

        df = pd.DataFrame(worksheet.values)
        print (df)
        

        df.to_excel(file_path, index = False)

        df_first_row = df.iloc[0].values.tolist()
        
        df_rest_rows = df.iloc[1:].values.tolist()

        # getting active sheet
        # active_sheet = wb.active
        # print(active_sheet)

        # reading a cell
        # print(worksheet["A1"].value)

        # excel_data = list()
        # # iterating over the rows and
        # # getting value from each cell in row
        # for row in worksheet.iter_rows():
        #     row_data = list()
        #     for cell in row:
        #         row_data.append(str(cell.value))
        #         # print(cell.value)
        #     excel_data.append(row_data)

        last_modified =  time.ctime(os.path.getmtime(file_path))
        created = time.ctime(os.path.getctime(file_path))
        return render(request, 'myapp/index.html', {"df_first_row":df_first_row, 
                                                    "df_rest_rows":df_rest_rows,
                                                    "last_modified":last_modified, 
                                                    "created":created})









